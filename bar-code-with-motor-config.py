import cv2
import RPi.GPIO as GPIO
from pyzbar.pyzbar import decode
import openpyxl as op
from gpiozero import AngularServo
from time import sleep
cap=cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)
camera=True
while camera==True:
    success, frame=cap.read()
    for code in decode(frame):
        print(code.type)
        print(code.data.decode('utf-8'))
        wb=op.load_workbook('test.xlsx')
        sh=wb.active
        h=int(code.data.decode('utf-8'))
        for i in range(2,6):
            c=sh.cell(row=i, column=2)
            if(c.value==h):
                d=sh.cell(row=i,column=3)
                print(d.value)
                n=int(input("Do u want?"))
                if(n==1):
                    if(d.value <5):
                        print('Insufficient balance')
                        break
                    d.value=d.value-5
                    wb.save('test.xlsx')
                    print(d.value)
                    print("your balance is ",d.value)
                    #DC motor working code
                    servo=AngularServo(26, min_angle=0, max_angle=270, min_pulse_width=0.0005, max_pulse_width=0.0025)
                    servo.angle=0
                    sleep(2)
                    servo.detach()
                else:
                    break
    cv2.imshow('Testing-code-scan',frame)
    cv2.waitKey(1)